import openpyxl
from datetime import date

wb = openpyxl.load_workbook("g:/mój dysk/grafik/grafiks.xlsx")
ms = input("podaj nr miesiaca: ")
sheet = wb[ms]


# ms = input("podaj numer miesiaca")
dt = date(2022, int(ms), 1)
if ms == "12":
    dt2 = date(2023, 1, 1)
if ms != "12":
    dt2 = date(2022, int(ms) + 1, 1)
days_in_month = abs(dt2 - dt)
print(days_in_month.days)
# days = (x + 1 for x in range(int(days_in_month.days)))
# for d in days:
#     cell = sheet.cell(row=int(d)+1, column=1)
#     cell.value = f"{d}. {ms}, 2022"
days = input("podaj dni dziennej zmiany odzielone spacja: ")
ndays = input("podaj dni nocnej zmiany odzielone spacja: ")
dayshift = days.split(' ')
nightshift = ndays.split(' ')
cell_suma = days_in_month.days + 2
for d in dayshift:
    cell = sheet.cell(row=int(d)+1, column=2)
    cell.value = "07:00"
for d in nightshift:
    cell = sheet.cell(row=int(d)+1, column=2)
    cell.value = "19:00"
godzin = (len(dayshift) + len(nightshift))*12
cell = sheet[f'b{cell_suma}']
cell.value = godzin


# cell = sheet['a1']
# cell.value = '123'
wb.save(f'g:/mój dysk/grafik/grafiks.xlsx')
